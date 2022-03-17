import xlrd
import openpyxl

if __name__ == "__main__":
    data = xlrd.open_workbook('/Users/Lenovo/Desktop/whnfull.xlsx')
    table = data.sheets()[0]
    nrows = table.nrows
    tmp = dict()
    result = dict()
    for i in range(nrows):
        if i == 0:
            continue
        row = table.row_values(i)
        key = row[0] + ':' + row[1]
        if key in tmp:
            if float(row[10]) < tmp.get(key):
                tmp[key] = float(row[10])
                result[key] = i
        else:
            tmp[key] = float(row[10])
            result[key] = i

    workbook = openpyxl.Workbook()
    result_sheet = workbook.create_sheet(index=0)
    tem = 1
    for key in result.keys():
        t_row = table.row_values(result[key])
        for i, content in enumerate(t_row):
            result_sheet.cell(tem, i+1).value = content
        tem = tem + 1

    workbook.save('/Users/Lenovo/Desktop/whnfull（无重复）.xlsx')
