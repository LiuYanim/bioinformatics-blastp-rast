import xlrd
import openpyxl

if __name__ == "__main__":
    data = xlrd.open_workbook('/Users/Lenovo/Desktop/whnfull.xlsx')
    table = data.sheets()[0]
    nrows = table.nrows
    tmp = dict()
    for i in range(nrows):
        if i == 0:
            continue
        row = table.row_values(i)
        key = row[0]
        if key in tmp:
            tmp[key].append(row[1])
        else:
            tmp[key] = [row[1]]

    workbook = openpyxl.Workbook()
    result_sheet = workbook.create_sheet(index=0)
    tem = 1
    for key in tmp.keys():
        result_sheet.cell(tem, 1).value = key
        for i, content in enumerate(tmp[key]):
            result_sheet.cell(tem, i + 1).value = content
        result_sheet.cell(tem, len(tmp[key]) + 2).value = len(tmp[key])
        tem = tem + 1

    workbook.save('/home/xiaoju/result.xls')
