import os
import re
import fileinput
import xlrd
import openpyxl

if __name__ == "__main__":
    primaries = list()
    products = list()
    try:
        for line in fileinput.input("/Users/didi/Downloads/rast.txt"):
            primary = re.search("/db_xref=(\"SEED:fig.*)", line)
            if primary is not None:
                print(primary.group(1)[1:-1])
                primaries.append(primary.group(1)[1:-1])
            product = re.search("/product=(.*)", line)
            if product is not None:
                print(product.group(1)[1:-1])
                products.append(product.group(1)[1:-1])
    except Exception as e:
        print("erro ..."+str(e))
    finally:
        print(len(primaries))
        print(len(products))
        print("finally ...")

    workbook = openpyxl.Workbook()
    result_sheet = workbook.create_sheet(index=0)
    data = xlrd.open_workbook('/home/xiaoju/data1.xls')
    table = data.sheets()[0]
    nrows = table.nrows
    for i in range(nrows):
        break_flag = False
        row = table.row_values(i)
        result_sheet.cell(i + 1, 1).value = row[0]
        result_sheet.cell(i + 1, 2).value = row[1]
        result_sheet.cell(i + 1, 3).value = row[2]
        first = row[0]
        if first not in primaries:
            continue
        if products[primaries.index(first)] != "hypothetical protein":
            result_sheet.cell(i + 1, 4).value = products[primaries.index(first)]
            continue
        else:
            for item in row[1].split(","):
                if item not in primaries:
                    continue
                if products[primaries.index(item)] != "hypothetical protein":
                    result_sheet.cell(i + 1, 4).value = products[primaries.index(item)]
                    break_flag = True
                    break

        if break_flag:
            continue
        else:
            result_sheet.cell(i + 1, 4).value = "unknown"

    workbook.save('/home/xiaoju/result.xls')
